from pkgutil import get_data
from pydoc import cli
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
import win32com.client
#from win32com.client import Dispatch
from .db import get_agente, get_productos, get_estadisticaCliente, get_saldoCliente, get_compPago, get_compVenta, cierreDeVenta
import os
import pythoncom
import winshell
from datetime import date, datetime, timedelta

escritorio = winshell.desktop()
hoy = date.today()
hoyf = hoy.strftime("%d%m%Y")
ruta = os.getcwd()

def catalogoAgentes():
    data = get_agente()
    archivo = "catalogoAgentes"
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    archivor = escritorio+"\\REPORTES\\"+archivo+"-"+hoyf
    archivoe = archivor+".xlsx"
    # Crear Excel
    wb = load_workbook(base)
    sheet = wb.active
    contador = 6
    contagent = 0
    agenteant = ""
    for i in range(len(data)):
        agente = data[contagent].get('id')
        nombre = data[contagent].get('name')
        correo = data[contagent].get('email')
        telefono = data[contagent].get('number')
        cliente = data[contagent].get('client')
        clientEmail = data[contagent].get('clientEmail')
        clientNumber = data[contagent].get('clientNumber')
        if agente != agenteant:
            sheet['A'+str(contador)] = nombre
            sheet['B'+str(contador)] = correo
            sheet['C'+str(contador)] = telefono
            sheet['D'+str(contador)] = cliente
            sheet['E'+str(contador)] = clientEmail
            sheet['F'+str(contador)] = clientNumber
        else:
            sheet['D'+str(contador)] = cliente
            sheet['E'+str(contador)] = clientEmail
            sheet['F'+str(contador)] = clientNumber
        agenteant = agente
        contador += 1 
        contagent += 1
    wb.save(archivoe) 
    crearPdf(archivor, archivo) 

def catalogoProductos():
    data = get_productos()
    archivo = "catalogoProductos"
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    archivor = escritorio+"\\REPORTES\\"+archivo+"-"+hoyf
    archivoe = archivor+".xlsx"
    # Crear Excel
    wb = load_workbook(base)
    sheet = wb.active
    contador = 5
    for i in range(len(data)):
        clave = data[i].get('nom_corto')
        descripcion = data[i].get('descripcion')
        nombre = data[i].get('nombre')
        existencia = data[i].get('existencia_real')
        sheet['A'+str(contador)] = clave
        sheet['B'+str(contador)] = descripcion
        sheet['C'+str(contador)] = nombre
        sheet['D'+str(contador)] = existencia 
        contador += 1 
    wb.save(archivoe) 
    crearPdf(archivor, archivo)       

def estadisticasCliente(id):
    data = get_estadisticaCliente(id)
    archivo = "estadisticasCliente"
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    archivor = escritorio+"\\REPORTES\\"+archivo+"-"+hoyf
    archivoe = archivor+".xlsx"
    # Crear Excel
    wb = load_workbook(base)
    sheet = wb.active
    cliente = data.get('cliente')
    agente = data.get('agente')
    ventaInfo = data.get('datosVenta')
    sheet['B'+str(3)] = cliente
    sheet['H'+str(3)] = agente
    contador = 6
    conta = 0
    ventaant = ""
    for i in range(len(ventaInfo)):
        vclave = ventaInfo[conta].get('vclave')
        vfecha = ventaInfo[conta].get('vfecha')
        vtotal = ventaInfo[conta].get('vtotal')
        vpago = ventaInfo[conta].get('vpago')
        pdescripcion = ventaInfo[conta].get('pdescripcion')
        pclave = ventaInfo[conta].get('pclave')
        pcantidad = ventaInfo[conta].get('pcantidad')
        pprecio = ventaInfo[conta].get('pprecio')
        pagopen = vtotal - vpago
        d0 = datetime.strptime(vfecha, "%Y-%m-%d")
        d1 = datetime.strptime(hoyf, "%d%m%Y")
        dvencidos = (d1 - d0).days
        if dvencidos < 30 : 
            dvencidos = 0
        else : 
            dvencidos = dvencidos - 30
        if vclave != ventaant:
            sheet['A'+str(contador)] = vclave 
            sheet['B'+str(contador)] = vfecha 
            sheet['C'+str(contador)] = vtotal
            sheet['D'+str(contador)] = pagopen
            sheet['E'+str(contador)] = dvencidos
            sheet['F'+str(contador)] = pdescripcion
            sheet['G'+str(contador)] = pclave
            sheet['H'+str(contador)] = pcantidad
            sheet['I'+str(contador)] = pprecio
        else:
            sheet['F'+str(contador)] = pdescripcion
            sheet['G'+str(contador)] = pclave
            sheet['H'+str(contador)] = pcantidad
            sheet['I'+str(contador)] = pprecio
        ventaant = vclave
        contador += 1 
        conta += 1
    wb.save(archivoe) 
    crearPdf(archivor, archivo)

def saldosCliente():
    information = get_saldoCliente()
    data = information['data']
    agentSum = information['agentSum']
    archivo = "saldoDeudorCliente"
    archivor = escritorio+"\\REPORTES\\"+archivo+"-"+hoyf
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    archivoe = archivor+".xlsx"
    # Crear Excel
    wb = load_workbook(base)
    sheet = wb.active
    head = 5
    for agent in data:
        sheet['A'+str(head)] = agent
        sheet['G'+str(head)] = agentSum[agent]
        sheet['G'+str(head)].font = Font(bold=True,color = "FFFF0000")
        head+=1
        for client in data[agent]:
            sheet['B'+str(head)] = client 
            head+=1
            for venta in data[agent][client]['data']:
                sheet['B'+str(head)] = venta['product']
                sheet['C'+str(head)] = venta['idVenta']
                sheet['D'+str(head)] = venta['fecha']
                sheet['E'+str(head)] = venta['totalPagar']
                sheet['F'+str(head)] = venta['montoPagado']
                sheet['G'+str(head)] = venta['debt']
                d0 = datetime.strptime(venta['fecha'], "%Y-%m-%d")
                d1 = datetime.strptime(hoyf, "%d%m%Y")
                distance = (d1 - d0).days
                sheet['H'+str(head)] = 0 if distance < 30 else distance - 30
                head+=1
            sheet['G'+str(head)]= data[agent][client]['debt']
            sheet['G'+str(head)].font = Font(bold=True,color = "FFFF0000")
            head+=1    
        head+=1 
    wb.save(archivoe) 
    crearPdf(archivor, archivo)

def comprobantePago(idVenta):
    data = get_compPago(idVenta)
    archivo = "comprobantePago"
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    cliente = data.get('cliente')
    agente = data.get('agente')
    vfecha = data.get('vfecha')
    vtotal = data.get('vtotal')
    vpagado = data.get('vpagado')
    historial = data.get('historial')
    fecha = datetime.strptime(vfecha, "%Y-%m-%d")
    flimite = fecha + timedelta(days=30)
    # Crear Excel
    archivor = escritorio+"\\REPORTES\\"+archivo+str(idVenta)+"-"+hoyf
    archivoe = archivor+".xlsx"
    wb = load_workbook(base)
    sheet = wb.active
    sheet['B'+str(6)] = cliente
    sheet['B'+str(7)] = agente
    sheet['F'+str(6)] = idVenta
    sheet['F'+str(7)] = fecha.strftime("%d/%m/%Y")
    sheet['F'+str(8)] = flimite
    sheet['F'+str(9)] = vtotal
    sheet['F'+str(10)] = vpagado
    filae = 13
    for i in range(len(historial)):
        fechap = historial[i].get('fecha')
        fechaf = datetime.strptime(vfecha, "%Y-%m-%d")
        montop = historial[i].get('monto')
        sheet['A'+str(filae)] = fechaf.strftime("%d/%m/%Y")
        sheet['B'+str(filae)] = montop
        filae += 1
    wb.save(archivoe) 
    crearPdf(archivor, archivo)  

def comprobanteVenta(idVenta):
    data = get_compVenta(idVenta)
    archivo = "comprobanteVenta"
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    cliente = data.get('cliente')
    agente = data.get('agente')
    vfecha = data.get('vfecha')
    vtotal = data.get('vtotal')
    productos = data.get('productos')
    fecha = datetime.strptime(vfecha, "%Y-%m-%d")
    # Crear Excel
    archivor = escritorio+"\\REPORTES\\"+archivo+str(idVenta)+"-"+hoyf
    archivoe = archivor+".xlsx"
    wb = load_workbook(base)
    sheet = wb.active
    sheet['B'+str(6)] = cliente
    sheet['B'+str(7)] = agente
    sheet['E'+str(6)] = idVenta
    sheet['E'+str(7)] = fecha.strftime("%d/%m/%Y")
    sheet['E'+str(8)] = vtotal
    filae = 11
    for i in range(len(productos)):
        codProducto = productos[i].get('idproducto')
        cantidad = productos[i].get('cantidad')
        preciou = productos[i].get('precio')
        descripcion = productos[i].get('descripcion')
        preciott = (preciou*cantidad)
        sheet['A'+str(filae)] = cantidad
        sheet['B'+str(filae)] = descripcion
        sheet['C'+str(filae)] = codProducto
        sheet['D'+str(filae)] = preciou
        sheet['E'+str(filae)] = preciott
        
        filae += 1
    wb.save(archivoe) 
    crearPdf(archivor, archivo)

def reporteCierreVenta(startDate, endDate=None):
    complete = cierreDeVenta(startDate, endDate)
    data = complete['info']
    productSum = complete['productSum']
    sumSell = complete['sumSell']

    archivo = "cierreVenta"
    archivor = escritorio+"\\REPORTES\\"+archivo+"-"+hoyf
    base = ruta+"\\reportes\\base\\o"+archivo+".xlsx"
    archivoe = archivor+".xlsx"
    # Crear Excel
    wb = load_workbook(base)
    sheet = wb.active    
    filae = 6
    finicio = datetime.strptime(startDate, "%Y%m%d")
    ffin = datetime.strptime(endDate, "%Y%m%d") if endDate is not None else ''
    sheet['H'+str(2)] = finicio.strftime("%d/%m/%Y")
    sheet['I'+str(2)] = "-"+str(ffin.strftime("%d/%m/%Y")) if endDate is not None else ''
    for i in range(len(data)):
        fechav = datetime.strptime(data[i].get('date'), "%Y-%m-%d")
        ventaId = data[i].get('ventaId')
        ventaFactura = data[i].get('ventaFactura')
        ventaCliente = data[i].get('ventaCliente')
        ventaTotal = data[i].get('total')
        pagado = data[i].get('pagado')
        if pagado < ventaTotal:
            credito = "Si"
        else:
            credito = "No"
        listProduct = data[i].get('listProduct')
        if ventaFactura == "" :
            tipo = "Remision"
        else:
            tipo = "Factura"
        sheet['A'+str(filae)] = fechav.strftime("%d/%m/%Y")
        sheet['B'+str(filae)] = tipo
        sheet['C'+str(filae)] = credito
        sheet['D'+str(filae)] = ventaCliente
        sheet['E'+str(filae)] = ventaId
        for producto in range(len(listProduct)):
            pclave = listProduct[producto].get("code")
            pdescripcion = listProduct[producto].get("name")
            cantidad = listProduct[producto].get("amount")
            pprecio = listProduct[producto].get("price")
            sheet['F'+str(filae)] = pclave
            sheet['G'+str(filae)] = pdescripcion
            sheet['H'+str(filae)] = cantidad
            sheet['I'+str(filae)] = pprecio
            filae += 1
        sheet['J'+str(filae)] = ventaTotal
        filae += 2
    
    sheet['G'+str(filae)] = "Total de ventas Efectivo"
    sheet['G'+str(filae)].font = Font(bold=True)
    sheet['I'+str(filae)] = sumSell['cash']
    filae += 1
    sheet['G'+str(filae)] = "Total de Credito"
    sheet['G'+str(filae)].font = Font(bold=True)
    sheet['I'+str(filae)] = sumSell['debt']

    

    filae += 3
    sheet['D'+str(filae)] = 'Producto'
    sheet['D'+str(filae)].font = Font(bold=True)
    filae += 1
    for product in productSum: 
        sheet['F'+str(filae)] = product
        sheet['H'+str(filae)] = productSum[product]['amount']
        sheet['I'+str(filae)] = productSum[product]['sum']
        filae+=1

    wb.save(archivoe) 
    crearPdf(archivor, archivo)

# Crear PDF
def crearPdf(archivor, archivo):
    try:
        pythoncom.CoInitialize()
        excel_file = win32com.client.Dispatch("Excel.Application")
        xl_sheets = excel_file.Workbooks.Open(archivor+".xlsx")
        worksheets = xl_sheets.Worksheets[0]
        worksheets.ExportAsFixedFormat(0, archivor+".pdf")
        excel_file.quit()
    except Exception as inst:
        print("OS error: {0}".format(inst))
    return
